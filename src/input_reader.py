"""Input file loading for CSV and Excel supplier files."""

from __future__ import annotations

from typing import Any, Optional
import polars as pl


def _seek_start(source: Any) -> None:
    if hasattr(source, "seek"):
        source.seek(0)


def _read_xlsx_openpyxl(source: Any, max_rows: Optional[int] = None) -> pl.DataFrame:
    from openpyxl import load_workbook

    _seek_start(source)
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pl.DataFrame()

    headers = []
    seen = {}
    for idx, raw in enumerate(rows[0], 1):
        base = str(raw).strip() if raw not in (None, "") else f"Column {idx}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")

    width = len(headers)
    data = []
    data_rows = rows[1:max_rows + 1] if max_rows else rows[1:]
    for row in data_rows:
        values = list(row[:width]) + [None] * max(0, width - len(row))
        data.append(values)
    return pl.DataFrame(data, schema=headers, orient="row")


def read_supplier_file(source: Any, filename: Optional[str] = None, max_rows: Optional[int] = None) -> pl.DataFrame:
    """Read a supplier CSV/XLSX file into a Polars DataFrame."""
    name = (filename or str(source)).lower()
    if name.endswith(".csv"):
        _seek_start(source)
        # Treat CSV client data as text so original columns round-trip without
        # type coercion such as TRUE/FALSE becoming false/true in the output.
        return pl.read_csv(source, n_rows=max_rows, infer_schema_length=0)
    if name.endswith((".xlsx", ".xlsm")):
        _seek_start(source)
        try:
            df = pl.read_excel(source)
            return df.head(max_rows) if max_rows else df
        except Exception:
            return _read_xlsx_openpyxl(source, max_rows=max_rows)
    if name.endswith(".xls"):
        _seek_start(source)
        df = pl.read_excel(source)
        return df.head(max_rows) if max_rows else df
    raise ValueError("Input file must be CSV or Excel (.csv, .xlsx, .xlsm, .xls)")
